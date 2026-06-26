"""
serial_comm.py — Comunicação serial com o Arduino.

Responsabilidades:
  • Detecção automática da porta serial
  • Leitura de dados do sensor em thread separada
  • Conversão de coordenadas polares → cartesianas
  • Sinalização de fim de leitura via callback

Autor : Yan de Lima Pereira
Versão: 2.0
"""

import logging
import threading
import time
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

try:
    import serial
    import serial.tools.list_ports
    SERIAL_DISPONIVEL = True
except ImportError:
    SERIAL_DISPONIVEL = False

from config import Config

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Detecção e conexão
# ─────────────────────────────────────────────────────────────────────────────

def detectar_porta_arduino() -> str | None:
    """
    Detecta automaticamente a porta serial do Arduino.

    Estratégia:
      1. Busca por descrição conhecida (arduino / ch340 / usb serial)
      2. Se única porta disponível, assume que é o Arduino
      3. Caso contrário retorna None e loga lista encontrada

    Returns:
        str com nome da porta, ou None se não encontrada
    """
    if not SERIAL_DISPONIVEL:
        logger.error("pyserial não instalado — modo serial indisponível")
        return None

    portas = list(serial.tools.list_ports.comports())
    if not portas:
        logger.warning("Nenhuma porta serial detectada")
        return None

    for p in portas:
        descricao = p.description.lower()
        if any(t in descricao for t in ("arduino", "ch340", "usb serial")):
            logger.info("Arduino detectado em %s (%s)", p.device, p.description)
            return p.device

    if len(portas) == 1:
        logger.info("Única porta disponível: %s — assumindo Arduino", portas[0].device)
        return portas[0].device

    logger.warning(
        "Múltiplas portas encontradas, nenhuma identificada como Arduino: %s",
        [p.device for p in portas],
    )
    return None


def _processar_linha(linha: str) -> tuple[list[float], list[float]] | None:
    """
    Converte linha serial do Arduino em ponto 3D cartesiano e extrai dados brutos.
    Retorna: (dados_interface, dados_planilha)
    """
    try:
        campos = {}
        for parte in linha.split(","):
            parte = parte.strip()
            if ":" not in parte:
                continue
            chave, valor = parte.split(":", 1)
            valor_num = valor.strip().lower().replace("mm", "").replace("°", "").strip()
            campos[chave.strip().lower()] = float(valor_num)

        def _get(keys):
            for k in keys:
                if k in campos:
                    return campos[k]
            return None

        camada      = _get(["camada", "cam"]) or 0.0  # Padrão 0 caso não venha
        distancia   = _get(["distância", "distancia", "dist"])
        angulo_mesa = _get(["ângulo", "angulo", "ang"])
        altura_fuso = _get(["altura", "alt"])

        if distancia is None or angulo_mesa is None or altura_fuso is None:
            return None

        raio  = Config.DISTANCIA_SENSOR_MESA - distancia
        theta = np.radians(angulo_mesa)
        x     = raio * np.cos(theta)
        y     = raio * np.sin(theta)
        z     = altura_fuso

        ponto_ui = [raio, theta, x, y, z]
        ponto_planilha = [camada, angulo_mesa, distancia, altura_fuso, raio, theta, x, y, z]

        return ponto_ui, ponto_planilha

    except (ValueError, AttributeError):
        return None



ARQUIVO_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dados_scanner.xlsx")
CABECALHOS = [
    "camada_bruta", "angulo_bruto", "distancia_bruta", "altura_bruta",
    "raio_calc", "theta_calc", "x_calc", "y_calc", "z_calc"
]

def _cabecalho_estilo(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="2F5496")
    cell.alignment = Alignment(horizontal="center")

def salvar_ponto_na_planilha(
    camada: float, angulo: float, distancia: float, altura: float,
    raio: float, theta: float, x: float, y: float, z: float,
    arquivo: str = ARQUIVO_XLSX, novo_lote: bool = False,
) -> None:

    arquivo_existe = os.path.exists(arquivo)

    if not arquivo_existe:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pontos 3D"
        for col, cab in enumerate(CABECALHOS, start=1):
            cell = ws.cell(row=1, column=col, value=cab)
            _cabecalho_estilo(cell)
            ws.column_dimensions[cell.column_letter].width = 15
        proxima_linha = 2
    else:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active

        ultima_linha = ws.max_row
        # Ajustado para checar as 9 colunas
        while ultima_linha > 1 and all(
            ws.cell(row=ultima_linha, column=c).value is None
            for c in range(1, 10)
        ):
            ultima_linha -= 1

        if novo_lote:
            inicio_lote = ultima_linha + 3
            for col, cab in enumerate(CABECALHOS, start=1):
                cell = ws.cell(row=inicio_lote, column=col, value=cab)
                _cabecalho_estilo(cell)
            proxima_linha = inicio_lote + 1
        else:
            proxima_linha = ultima_linha + 1

    ws.cell(row=proxima_linha, column=1, value=camada)
    ws.cell(row=proxima_linha, column=2, value=angulo)
    ws.cell(row=proxima_linha, column=3, value=distancia)
    ws.cell(row=proxima_linha, column=4, value=altura)
    ws.cell(row=proxima_linha, column=5, value=round(raio, 4))
    ws.cell(row=proxima_linha, column=6, value=round(theta, 6))
    ws.cell(row=proxima_linha, column=7, value=round(x, 4))
    ws.cell(row=proxima_linha, column=8, value=round(y, 4))
    ws.cell(row=proxima_linha, column=9, value=round(z, 4))

    wb.save(arquivo)

"""
import random
import math

# Simula saídas do _processar_linha já processadas
def gerar_ponto_falso():
    raio  = random.uniform(50, 150)
    theta = random.uniform(0, 2 * math.pi)
    x     = raio * math.cos(theta)
    y     = raio * math.sin(theta)
    z     = random.uniform(0, 100)
    return [raio, theta, x, y, z]

# Simula 2 sessões de scan com 5 pontos cada
for sessao in range(2):
    print(f"\n--- Sessão {sessao + 1} ---")
    for i in range(5):
        ponto = gerar_ponto_falso()
        salvar_ponto_na_planilha(*ponto, novo_lote=(i == 0 and sessao > 0))
        print(f"Ponto salvo: {[round(v, 3) for v in ponto]}")
"""

# ─────────────────────────────────────────────────────────────────────────────
# Leitura em background
# ─────────────────────────────────────────────────────────────────────────────

def ler_dados_async(
    callback_pontos,
    callback_erro=None,
    callback_ponto_novo=None,
    porta: str | None = None,
    baudrate: int | None = None,
    timeout: float | None = None,
    stop_event: threading.Event | None = None,
) -> tuple[threading.Thread, threading.Event]:
    """
    Inicia leitura serial em thread daemon para não bloquear a UI.

    Args:
        callback_pontos : ``f(np.ndarray)`` — chamado ao final da leitura
        callback_erro   : ``f(str)``        — chamado em caso de falha serial
        callback_ponto_novo: ``f([x,y,z])`` — chamado a cada ponto recebido
        porta           : Porta serial (auto-detecta se None)
        baudrate        : Taxa de comunicação (padrão: Config.BAUDRATE_SERIAL)
        timeout         : Timeout de leitura em segundos
        stop_event      : Event externo para cancelar a leitura; se None, cria um novo

    Returns:
        Tupla (thread, stop_event) — chame stop_event.set() para interromper
    """
    if baudrate is None:
        baudrate = Config.BAUDRATE_SERIAL
    if timeout is None:
        timeout = Config.TIMEOUT_SERIAL
    if stop_event is None:
        stop_event = threading.Event()

    def _tarefa():
        print("[SERIAL] thread iniciado", flush=True)
        pontos = []
        primeiro_ponto_do_lote = True
        porta_alvo = porta or detectar_porta_arduino()

        if not porta_alvo:
            msg = "Porta do Arduino não encontrada. Verifique a conexão."
            logger.error(msg)
            if callback_erro:
                callback_erro(msg)
            return

        print(f"[SERIAL] porta detectada: {porta_alvo}", flush=True)

        try:
            print(f"[SERIAL] conectando em {porta_alvo} @ {baudrate} baud…", flush=True)
            logger.info("Conectando em %s @ %d baud…", porta_alvo, baudrate)
            with serial.Serial(porta_alvo, baudrate, timeout=timeout) as ser:
                time.sleep(2)   # Aguarda reset do Arduino
                print("[SERIAL] conexão estabelecida — aguardando dados…", flush=True)
                logger.info("Conexão estabelecida — aguardando dados…")

                linhas_recebidas = 0
                while not stop_event.is_set():
                    linha = ser.readline().decode(errors="ignore").strip()

                    if not linha:
                        # timeout sem dados — imprime um ponto a cada 10 timeouts para confirmar que o loop está vivo
                        linhas_recebidas += 1
                        if linhas_recebidas % 10 == 0:
                            print(f"[SERIAL] aguardando... ({linhas_recebidas} timeouts, {len(pontos)} pontos até agora)", flush=True)
                        continue

                    print(f"[SERIAL] linha bruta recebida: {repr(linha)}", flush=True)

                    if linha.upper() == "FIM":
                        print("[SERIAL] sinal FIM recebido — encerrando leitura", flush=True)
                        logger.info("Sinal FIM recebido — leitura concluída")
                        break

                    dados_processados = _processar_linha(linha)
                    if dados_processados is None:
                        print(f"[SERIAL] linha ignorada (formato inválido): {repr(linha)}", flush=True)
                        continue

                    ponto_ui, ponto_planilha = dados_processados

                    pontos.append(ponto_ui)
                    x, y, z = ponto_ui[2], ponto_ui[3], ponto_ui[4]
                    n = len(pontos)
                    print(f"[SERIAL] #{n:>4}  x={x:8.2f}  y={y:8.2f}  z={z:8.2f}  mm", flush=True)
                    logger.debug("Ponto #%d: x=%.2f y=%.2f z=%.2f", n, x, y, z)

                    try:
                        salvar_ponto_na_planilha(*ponto_planilha, novo_lote=primeiro_ponto_do_lote)
                        primeiro_ponto_do_lote = False   # só avança após salvar com sucesso
                    except PermissionError as e:
                        msg = (
                            f"Sem permissão para gravar a planilha.\n"
                            f"Se '{os.path.basename(ARQUIVO_XLSX)}' estiver aberto no Excel, feche-o e tente novamente.\n"
                            f"Detalhe: {e}"
                        )
                        logger.error(msg)
                        print(f"[PLANILHA] PermissionError — {e}", flush=True)
                        if callback_erro:
                            callback_erro(msg)
                        return   # encerra o thread — continuar sobrescreveria dados corrompidos
                    except Exception as e:
                        logger.warning("Falha ao salvar ponto na planilha: %s", e)
                        print(f"[PLANILHA] erro ao salvar: {e}", flush=True)
                        if callback_erro:
                            callback_erro(f"Erro ao salvar na planilha: {e}")

                    if callback_ponto_novo:
                        callback_ponto_novo(ponto_ui[2:5])

                if stop_event.is_set():
                    print("[SERIAL] leitura interrompida externamente", flush=True)
                    logger.info("Leitura interrompida externamente — porta fechada")
                    return

        except serial.SerialException as e:
            msg = f"Erro na porta {porta_alvo}: {e}"
            print(f"[SERIAL] SerialException — {msg}", flush=True)
            logger.error(msg)
            if callback_erro:
                callback_erro(msg)
            return
        except Exception as e:
            # captura qualquer outra exceção que mataria o thread silenciosamente
            print(f"[SERIAL] EXCEÇÃO NÃO ESPERADA — {type(e).__name__}: {e}", flush=True)
            logger.exception("Exceção inesperada no thread serial")
            if callback_erro:
                callback_erro(f"Erro inesperado: {e}")
            return

        pts_arr = np.array(pontos) if pontos else np.empty((0, 5))
        print(f"[SERIAL] leitura finalizada — {len(pts_arr)} pontos capturados", flush=True)
        logger.info("Leitura serial finalizada — %d pontos capturados", len(pts_arr))
        callback_pontos(pts_arr)

    t = threading.Thread(target=_tarefa, daemon=True)
    t.start()
    return t, stop_event